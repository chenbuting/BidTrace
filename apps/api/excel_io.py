# -*- coding: utf-8 -*-
"""Excel 导入导出：表头必须与固定模板完全一致，否则拒绝导入。"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook

# 平台账号：中文表头 -> 字段（顺序即固定模板顺序）
PLATFORM_HEADERS = [
    ("平台名称", "name"),
    ("平台网址", "url"),
    ("登录方式", "login_method"),
    ("登录账号", "login_account"),
    ("登录密码", "login_password"),
    ("是否有CA证书", "has_ca"),
    ("CA证书密码", "ca_password"),
    ("平台优先级", "priority"),
    ("平台状态", "status"),
    ("平台权重(0~5)", "weight"),
    ("备注", "remark"),
]

# 询标：中文表头 -> 字段（顺序即固定模板顺序）
INQUIRY_HEADERS = [
    ("报名时间", "register_date"),
    ("平台", "platform_name"),
    ("项目名", "project_name"),
    ("是否投标", "is_bid"),
    ("是否报名", "is_registered"),
    ("文件是否领取", "file_received"),
    ("是否交费", "is_paid"),
    ("概况是否完成", "overview_done"),
    ("未参与原因类别", "skip_reason_category"),
    ("参与状态或未参与详细原因", "skip_reason_detail"),
    ("报名截止时间", "deadline"),
]

# 投标项目：中文表头 -> 字段（顺序即固定模板顺序）
BID_PROJECT_HEADERS = [
    ("序号", "serial_no"),
    ("开标时间", "open_time"),
    ("投标员", "bidder"),
    ("项目名称", "project_name"),
    ("平台", "platform"),
    ("备注", "remark"),
    ("是/否中标", "is_won"),
    ("中标金额", "win_amount"),
    ("是/否废标", "is_void"),
    ("投标金额", "bid_amount"),
    ("付款方式", "payment_method"),
]

# 投标保证金：第 6 列表头必须是字面量「金额\n（万元）」
DEPOSIT_HEADERS = [
    ("序号", "serial_no"),
    ("申请时间", "apply_time"),
    ("项目名称", "project_name"),
    ("收款单位", "payee"),
    ("平台", "platform"),
    ("金额\n（万元）", "amount"),
    ("投标员", "bidder"),
    ("是否退回", "is_returned"),
    ("保证金退回联系方式", "return_contact"),
    ("备注", "remark"),
]


class TemplateError(ValueError):
    """Excel 表头与固定模板不一致。"""


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _to_iso_date(raw: str) -> str:
    """统一成 YYYY-MM-DD，兼容 YYYYMMDD。"""
    s = (raw or "").strip()
    if not s:
        return ""
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        return s
    if len(s) == 8 and s.isdigit():
        return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"
    return s


def _normalize_cell_date(val: Any) -> str:
    """Excel 单元格日期/datetime 尽量规范为 YYYY-MM-DD。"""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return _to_iso_date(_cell_str(val))


def _read_header_row(ws: Any, expected_count: int) -> list[str]:
    """读取第 1 行表头；只取模板列数，后面若还有非空列也算违规。"""
    headers: list[str] = []
    max_c = max(ws.max_column or 1, expected_count)
    for c in range(1, max_c + 1):
        headers.append(_cell_str(ws.cell(1, c).value))
    return headers


def _validate_headers(actual: list[str], expected: list[tuple[str, str]], kind: str) -> None:
    """表头必须与模板列名、顺序完全一致；不允许缺列、错列、多列。"""
    expected_names = [zh for zh, _ in expected]
    n = len(expected_names)

    # 前面 n 列必须一模一样
    actual_n = actual[:n]
    while len(actual_n) < n:
        actual_n.append("")

    mismatches: list[str] = []
    for i, (got, want) in enumerate(zip(actual_n, expected_names), start=1):
        if got != want:
            mismatches.append(f"第{i}列应为「{want}」，实际为「{got or '（空）'}」")

    # 模板列之后不允许再有非空表头
    extras = [h for h in actual[n:] if h]
    if extras:
        mismatches.append(f"多出了未允许的列：{'、'.join(extras)}")

    if mismatches:
        tpl = "、".join(expected_names)
        raise TemplateError(
            f"{kind}表头与固定模板不一致，已拒绝导入。"
            f"正确表头（按顺序）为：{tpl}。"
            f"问题：{'；'.join(mismatches)}"
        )


def parse_platforms_xlsx(content: bytes) -> list[dict[str, Any]]:
    """解析平台账号 Excel（表头必须完全匹配模板）。"""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    try:
        ws = wb.active
        headers = _read_header_row(ws, len(PLATFORM_HEADERS))
        _validate_headers(headers, PLATFORM_HEADERS, "平台账号")

        rows: list[dict[str, Any]] = []
        for r in range(2, (ws.max_row or 1) + 1):
            item: dict[str, Any] = {f: ("" if f != "weight" else 0) for _, f in PLATFORM_HEADERS}
            empty = True
            for c, (_, field) in enumerate(PLATFORM_HEADERS, start=1):
                val = ws.cell(r, c).value
                if val is not None and str(val).strip() != "":
                    empty = False
                if field == "weight":
                    try:
                        item[field] = float(val) if val is not None and str(val).strip() != "" else 0
                    except (TypeError, ValueError):
                        item[field] = 0
                else:
                    item[field] = _cell_str(val)
            if not empty and item.get("name"):
                rows.append(item)
        return rows
    finally:
        wb.close()


def export_platforms_xlsx(rows: list[dict[str, Any]], mask_password: bool = True) -> bytes:
    """导出平台账号 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "投标平台账号数据"
    for i, (zh, _) in enumerate(PLATFORM_HEADERS, start=1):
        ws.cell(1, i, zh)
    for ri, row in enumerate(rows, start=2):
        for ci, (_, field) in enumerate(PLATFORM_HEADERS, start=1):
            val = row.get(field, "")
            if mask_password and field in ("login_password", "ca_password") and val:
                val = "***"
            ws.cell(ri, ci, val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def empty_platforms_template_xlsx() -> bytes:
    """空模板（仅表头）。"""
    return export_platforms_xlsx([], mask_password=False)


def parse_inquiries_xlsx(content: bytes) -> list[dict[str, Any]]:
    """解析询标 Excel（表头必须完全匹配模板）。"""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    try:
        ws = wb.active
        headers = _read_header_row(ws, len(INQUIRY_HEADERS))
        _validate_headers(headers, INQUIRY_HEADERS, "询标报名")

        rows: list[dict[str, Any]] = []
        for r in range(2, (ws.max_row or 1) + 1):
            item: dict[str, Any] = {f: "" for _, f in INQUIRY_HEADERS}
            empty = True
            for c, (_, field) in enumerate(INQUIRY_HEADERS, start=1):
                val = _cell_str(ws.cell(r, c).value)
                if val:
                    empty = False
                item[field] = val
            if not empty and (item.get("project_name") or item.get("platform_name")):
                item["register_date"] = _to_iso_date(str(item.get("register_date") or ""))
                item["deadline"] = _to_iso_date(str(item.get("deadline") or ""))
                rows.append(item)
        return rows
    finally:
        wb.close()


def export_inquiries_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """导出询标 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "询标报名跟踪"
    for i, (zh, _) in enumerate(INQUIRY_HEADERS, start=1):
        ws.cell(1, i, zh)
    for ri, row in enumerate(rows, start=2):
        for ci, (_, field) in enumerate(INQUIRY_HEADERS, start=1):
            ws.cell(ri, ci, row.get(field, ""))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def empty_inquiries_template_xlsx() -> bytes:
    """空模板（仅表头）。"""
    return export_inquiries_xlsx([])


def parse_bid_projects_xlsx(content: bytes) -> list[dict[str, Any]]:
    """解析投标项目 Excel（表头必须完全匹配模板）。"""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    try:
        ws = wb.active
        headers = _read_header_row(ws, len(BID_PROJECT_HEADERS))
        _validate_headers(headers, BID_PROJECT_HEADERS, "投标项目")

        rows: list[dict[str, Any]] = []
        for r in range(2, (ws.max_row or 1) + 1):
            item: dict[str, Any] = {f: "" for _, f in BID_PROJECT_HEADERS}
            empty = True
            for c, (_, field) in enumerate(BID_PROJECT_HEADERS, start=1):
                raw = ws.cell(r, c).value
                if raw is not None and str(raw).strip() != "":
                    empty = False
                if field == "open_time":
                    item[field] = _normalize_cell_date(raw)
                else:
                    item[field] = _cell_str(raw)
            if not empty and item.get("project_name"):
                rows.append(item)
        return rows
    finally:
        wb.close()


def export_bid_projects_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """导出投标项目 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "投标项目"
    for i, (zh, _) in enumerate(BID_PROJECT_HEADERS, start=1):
        ws.cell(1, i, zh)
    for ri, row in enumerate(rows, start=2):
        for ci, (_, field) in enumerate(BID_PROJECT_HEADERS, start=1):
            ws.cell(ri, ci, row.get(field, ""))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def empty_bid_projects_template_xlsx() -> bytes:
    """投标项目空模板（仅表头）。"""
    return export_bid_projects_xlsx([])


def parse_deposits_xlsx(content: bytes) -> list[dict[str, Any]]:
    """解析投标保证金 Excel（表头必须完全匹配模板）。"""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    try:
        ws = wb.active
        headers = _read_header_row(ws, len(DEPOSIT_HEADERS))
        _validate_headers(headers, DEPOSIT_HEADERS, "投标保证金")

        rows: list[dict[str, Any]] = []
        for r in range(2, (ws.max_row or 1) + 1):
            item: dict[str, Any] = {f: "" for _, f in DEPOSIT_HEADERS}
            empty = True
            for c, (_, field) in enumerate(DEPOSIT_HEADERS, start=1):
                raw = ws.cell(r, c).value
                if raw is not None and str(raw).strip() != "":
                    empty = False
                if field == "apply_time":
                    item[field] = _normalize_cell_date(raw)
                else:
                    item[field] = _cell_str(raw)
            if not empty and (item.get("project_name") or item.get("payee")):
                rows.append(item)
        return rows
    finally:
        wb.close()


def export_deposits_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """导出投标保证金 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "投标保证金"
    for i, (zh, _) in enumerate(DEPOSIT_HEADERS, start=1):
        ws.cell(1, i, zh)
    for ri, row in enumerate(rows, start=2):
        for ci, (_, field) in enumerate(DEPOSIT_HEADERS, start=1):
            ws.cell(ri, ci, row.get(field, ""))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def empty_deposits_template_xlsx() -> bytes:
    """投标保证金空模板（仅表头）。"""
    return export_deposits_xlsx([])


def _safe_sheet_title(name: str, used: set[str] | None = None) -> str:
    """Excel 工作表名：最多 31 字符，去掉非法字符，避免重名。"""
    raw = str(name or "未命名").strip() or "未命名"
    for ch in ("\\", "/", "?", "*", "[", "]", ":"):
        raw = raw.replace(ch, "")
    raw = raw[:31] or "未命名"
    used = used if used is not None else set()
    base = raw
    n = 2
    while raw in used:
        suffix = f"_{n}"
        raw = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(raw)
    return raw


def _fill_weekly_sheet(ws: Any, report: dict[str, Any]) -> None:
    """把一份周报写入已有工作表（样式对齐手工模板）。"""
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.styles.colors import Color

    # 与手工模板一致：细黑边框 + 分区浅蓝底（Office 主题色 Accent1 加亮）
    thin_side = Side(style="thin", color="FF000000")
    thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill(
        fill_type="solid",
        fgColor=Color(theme=4, tint=0.5999938962981048),
    )
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
    font_title = Font(name="宋体", size=20, bold=True, color="FF000000")
    font_meta = Font(name="宋体", size=11, bold=True, color="FF000000")
    font_section = Font(name="宋体", size=12, bold=True, color="FF000000")
    font_fallback = Font(name="宋体", size=11, bold=False, color="FF000000")
    # 条目富文本：编号标题加粗，正文不加粗
    inline_bold = InlineFont(rFont="宋体", sz=11, b=True, color="FF000000")
    inline_plain = InlineFont(rFont="宋体", sz=11, b=False, color="FF000000")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 列宽对齐模板
    widths = {
        "A": 8.44,
        "B": 4.46,
        "C": 12.0,
        "D": 0.91,
        "E": 11.0,
        "F": 19.82,
        "G": 41.73,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    def paint_row(r: int) -> None:
        for c in range(1, 8):
            ws.cell(r, c).border = thin

    def write_merged(
        r: int,
        value: Any,
        *,
        font: Font,
        fill: PatternFill | None,
        align: Alignment,
        height: float,
    ) -> None:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        cell = ws.cell(r, 1)
        cell.value = value
        # 富文本以 InlineFont 为准；普通字符串仍用 cell.font
        if not isinstance(value, CellRichText):
            cell.font = font
        cell.alignment = align
        if fill is not None:
            cell.fill = fill
        paint_row(r)
        ws.row_dimensions[r].height = height

    def format_item_rich(idx: int, title: str, body: str) -> tuple[Any, str]:
        """返回 (单元格值, 纯文本用于估高)：标题加粗，正文不加粗。"""
        title = (title or "").strip()
        body = (body or "").strip()
        head = f"{idx}.{title}" if title else f"{idx}."
        if not body:
            return CellRichText(TextBlock(inline_bold, head)), head
        plain = f"{head}\n\n{body}"
        rich = CellRichText(
            TextBlock(inline_bold, head),
            TextBlock(inline_plain, f"\n\n{body}"),
        )
        return rich, plain

    def estimate_height(text: str, base: float = 36.0) -> float:
        lines = max(1, text.count("\n") + 1)
        # 长行按约 28 字折行估算
        wrap_extra = 0
        for line in text.split("\n"):
            wrap_extra += max(0, (len(line) - 1) // 28)
        return max(base, 18.0 + 16.0 * (lines + wrap_extra))

    def write_items(items: list, *, empty_height: float, item_base: float) -> None:
        nonlocal row
        if not items:
            write_merged(
                row,
                "",
                font=font_fallback,
                fill=white_fill,
                align=align_left_center,
                height=empty_height,
            )
            row += 1
            return
        for i, it in enumerate(items, start=1):
            value, plain = format_item_rich(
                i, str(it.get("title") or ""), str(it.get("body") or "")
            )
            write_merged(
                row,
                value,
                font=font_fallback,
                fill=white_fill,
                align=align_left_center,
                height=estimate_height(plain, item_base),
            )
            row += 1

    display = str(report.get("display_name") or report.get("username") or "")
    week_label_text = str(report.get("week_label") or "")
    done_items = report.get("done_items") or []
    plan_items = report.get("plan_items") or []
    problem_items = report.get("problem_items") or []
    solution_items = report.get("solution_items") or []
    # 兼容旧字段
    if not problem_items and str(report.get("problems") or "").strip():
        problem_items = [{"title": "", "body": str(report.get("problems") or "").strip()}]
    if not solution_items and str(report.get("solutions") or "").strip():
        solution_items = [{"title": "", "body": str(report.get("solutions") or "").strip()}]

    # 第 1 行：标题
    write_merged(1, "工 作 报 表", font=font_title, fill=None, align=align_center, height=36.6)

    # 第 2 行：制表人 / 时间
    ws.merge_cells("A2:E2")
    ws.merge_cells("F2:G2")
    c_name = ws.cell(2, 1, f"制表人：{display}")
    c_name.font = font_meta
    c_name.alignment = align_left_center
    c_time = ws.cell(2, 6, f"时间：{week_label_text}")
    c_time.font = font_meta
    c_time.alignment = align_left_center
    paint_row(2)
    ws.row_dimensions[2].height = 31.2

    row = 3
    write_merged(row, "所做事项", font=font_section, fill=header_fill, align=align_center, height=31.0)
    row += 1
    write_items(done_items, empty_height=58.0, item_base=58.0)

    # 模板里事项后有一空白行
    write_merged(row, "", font=font_fallback, fill=white_fill, align=align_left_center, height=40.0)
    row += 1

    write_merged(row, "所遇问题", font=font_section, fill=header_fill, align=align_center, height=25.0)
    row += 1
    write_items(problem_items, empty_height=40.0, item_base=40.0)

    write_merged(row, "解决意见", font=font_section, fill=header_fill, align=align_center, height=25.0)
    row += 1
    write_items(solution_items, empty_height=40.0, item_base=40.0)

    write_merged(row, "预期工作", font=font_section, fill=header_fill, align=align_center, height=25.0)
    row += 1
    write_items(plan_items, empty_height=58.0, item_base=58.0)


def export_weekly_report_xlsx(report: dict[str, Any]) -> bytes:
    """按「工作报表」模板样式导出单份周报。"""
    wb = Workbook()
    ws = wb.active
    name = str(report.get("display_name") or report.get("username") or "周报")
    ws.title = _safe_sheet_title(name)
    _fill_weekly_sheet(ws, report)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_weekly_team_xlsx(reports: list[dict[str, Any]]) -> bytes:
    """合并多份已交周报：每人一个工作表（表名=姓名）。"""
    if not reports:
        raise ValueError("本周暂无已提交周报")
    wb = Workbook()
    # 先删掉默认空表，按人重建
    default = wb.active
    wb.remove(default)
    used: set[str] = set()
    for report in reports:
        name = str(report.get("display_name") or report.get("username") or "未命名")
        ws = wb.create_sheet(title=_safe_sheet_title(name, used))
        _fill_weekly_sheet(ws, report)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
