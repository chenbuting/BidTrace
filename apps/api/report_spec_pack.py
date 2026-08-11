# -*- coding: utf-8 -*-
"""报告规格辅助：规范化 AI 结果，并导出接近样例结构的 Excel。"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 导出配色：表头/必须改/建议/示例结果等，方便扫一眼
FILL_TITLE = PatternFill("solid", fgColor="FFF1EB")
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_LABEL = PatternFill("solid", fgColor="D6E3F0")
FILL_WARN = PatternFill("solid", fgColor="FFF3CD")
FILL_MUST = PatternFill("solid", fgColor="FCE8E6")
FILL_SUGGEST = PatternFill("solid", fgColor="E8F5E9")
FILL_NEW = PatternFill("solid", fgColor="C8E6C9")
FILL_DRAFT = PatternFill("solid", fgColor="FFF8E1")
FILL_PARAM = PatternFill("solid", fgColor="E3F2FD")
FILL_STEP = PatternFill("solid", fgColor="F5F5F5")
FILL_ALT = PatternFill("solid", fgColor="FAFAF8")
FILL_P = PatternFill("solid", fgColor="C8E6C9")
FILL_F = PatternFill("solid", fgColor="FFCDD2")
FILL_N = PatternFill("solid", fgColor="E0E0E0")

FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_TITLE = Font(bold=True, size=13, color="C2410C")
FONT_BOLD = Font(bold=True)
FONT_WARN = Font(bold=True, color="B54708")
ALIGN_TOP = Alignment(wrap_text=True, vertical="top")


def _s(v: Any) -> str:
    return str(v or "").strip()


def _list_of_dict(val: Any) -> list[dict[str, Any]]:
    if not isinstance(val, list):
        return []
    return [x for x in val if isinstance(x, dict)]


def normalize_report_spec_result(parsed: Any) -> dict[str, Any]:
    """把模型 JSON 整理成前端/导出可用的统一结构。"""
    data = parsed if isinstance(parsed, dict) else {}

    summary = _s(data.get("summary"))
    warnings: list[str] = []
    for w in data.get("warnings") or []:
        t = _s(w)
        if t:
            warnings.append(t)

    matches: list[dict[str, str]] = []
    for it in _list_of_dict(data.get("matches")):
        target = _s(it.get("target_spec"))
        if not target and not _s(it.get("base_report_no")):
            continue
        matches.append(
            {
                "target_spec": target,
                "base_report_no": _s(it.get("base_report_no")),
                "base_spec": _s(it.get("base_spec")),
                "reason": _s(it.get("reason")),
            }
        )

    changes: list[dict[str, str]] = []
    for it in _list_of_dict(data.get("changes")):
        position = _s(it.get("position") or it.get("field"))
        if not position and not _s(it.get("new_value")):
            continue
        changes.append(
            {
                "target_spec": _s(it.get("target_spec")),
                "position": position,
                "old_value": _s(it.get("old_value")),
                "new_value": _s(it.get("new_value")),
                "must_change": _s(it.get("must_change") or "建议"),
                "note": _s(it.get("note")),
            }
        )

    # 兼容旧版 items → 并入 changes
    for it in _list_of_dict(data.get("items")):
        position = _s(it.get("field") or it.get("position"))
        if not position and not _s(it.get("new_value")):
            continue
        changes.append(
            {
                "target_spec": _s(it.get("target_spec") or it.get("report_no")),
                "position": position,
                "old_value": _s(it.get("old_value")),
                "new_value": _s(it.get("new_value")),
                "must_change": _s(it.get("must_change") or "建议"),
                "note": _s(it.get("note")),
            }
        )

    test_items: list[dict[str, str]] = []
    for it in _list_of_dict(data.get("test_items")):
        item = _s(it.get("item"))
        if not item:
            continue
        test_items.append(
            {
                "target_spec": _s(it.get("target_spec")),
                "seq": _s(it.get("seq")),
                "item": item,
                "unit": _s(it.get("unit")),
                "requirement": _s(it.get("requirement")),
                "result_draft": _s(it.get("result_draft")),
                "rating": _s(it.get("rating")),
                "note": _s(it.get("note")),
            }
        )

    key_params: list[dict[str, str]] = []
    for it in _list_of_dict(data.get("key_params")):
        param = _s(it.get("param"))
        if not param:
            continue
        key_params.append(
            {
                "target_spec": _s(it.get("target_spec")),
                "param": param,
                "ref_value": _s(it.get("ref_value")),
                "note": _s(it.get("note")),
            }
        )

    steps: list[str] = []
    for s in data.get("steps") or []:
        t = _s(s)
        if t:
            steps.append(t)

    relative_diffs: list[dict[str, str]] = []
    for it in _list_of_dict(data.get("relative_diffs") or data.get("diff_highlights")):
        aspect = _s(it.get("aspect") or it.get("item") or it.get("field"))
        if not aspect and not _s(it.get("new_value")):
            continue
        relative_diffs.append(
            {
                "target_spec": _s(it.get("target_spec") or it.get("spec")),
                "aspect": aspect,
                "old_value": _s(it.get("old_value") or it.get("old")),
                "new_value": _s(it.get("new_value") or it.get("new")),
                "reason": _s(it.get("reason") or it.get("note")),
            }
        )

    # 兼容旧前端：扁平 items
    items = [
        {
            "report_no": c.get("target_spec") or "",
            "field": c.get("position") or "",
            "old_value": c.get("old_value") or "",
            "new_value": c.get("new_value") or "",
            "note": c.get("note") or "",
        }
        for c in changes
    ]

    return {
        "summary": summary,
        "warnings": warnings,
        "matches": matches,
        "relative_diffs": relative_diffs,
        "changes": changes,
        "test_items": test_items,
        "key_params": key_params,
        "steps": steps,
        "items": items,
    }


def _paint(cell: Any, fill: PatternFill | None = None, font: Font | None = None) -> None:
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    cell.alignment = ALIGN_TOP


def _is_must(text: str) -> bool:
    t = (text or "").strip()
    return "必须" in t and "非必须" not in t


def _rating_fill(rating: str) -> PatternFill | None:
    r = (rating or "").strip().upper()
    if r == "P":
        return FILL_P
    if r == "F":
        return FILL_F
    if r in {"N", "/"}:
        return FILL_N
    return None


def _write_sheet(
    ws: Any,
    headers: list[str],
    rows: list[list[Any]],
    *,
    start_row: int = 1,
    must_col: int | None = None,
    new_col: int | None = None,
    draft_col: int | None = None,
    rating_col: int | None = None,
    param_col: int | None = None,
) -> None:
    """写表头+数据行，并按列/必须改着色。"""
    header_row = start_row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(header_row, col, h)
        _paint(cell, FILL_HEADER, FONT_HEADER)
    for r_i, row in enumerate(rows):
        excel_row = header_row + 1 + r_i
        must = False
        if must_col and must_col <= len(row):
            must = _is_must(str(row[must_col - 1] or ""))
        row_fill = FILL_MUST if must else (FILL_ALT if excel_row % 2 == 0 else None)
        for c_i, val in enumerate(row, 1):
            cell = ws.cell(excel_row, c_i, val)
            fill = row_fill
            font = None
            if new_col and c_i == new_col:
                fill = FILL_NEW
            elif draft_col and c_i == draft_col:
                fill = FILL_DRAFT
            elif param_col and c_i == param_col:
                fill = FILL_PARAM
            elif must_col and c_i == must_col:
                fill = FILL_MUST if must else FILL_SUGGEST
                font = FONT_BOLD
            elif rating_col and c_i == rating_col:
                fill = _rating_fill(str(val)) or fill
            _paint(cell, fill, font)
    widths = [18, 22, 28, 28, 14, 28, 12, 24]
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = (
            widths[i - 1] if i - 1 < len(widths) else 18
        )


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Excel 工作表名合法化（最长31，去非法字符，避免重名）。"""
    bad = set(r"\/*?:[]")
    base = "".join(ch for ch in (name or "规格").strip() if ch not in bad) or "规格"
    base = base[:28]
    title = base
    i = 2
    while title in used:
        suffix = f"_{i}"
        title = (base[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(title)
    return title


def _ordered_specs(result: dict[str, Any]) -> list[str]:
    """按 matches → changes → test_items 顺序收集目标规格。"""
    specs: list[str] = []
    for src in (
        result.get("matches") or [],
        result.get("relative_diffs") or [],
        result.get("changes") or [],
        result.get("test_items") or [],
        result.get("key_params") or [],
    ):
        for it in src:
            s = _s(it.get("target_spec"))
            if s and s not in specs:
                specs.append(s)
    return specs


def _short_spec_label(spec: str, idx: int) -> str:
    """工作表短名：规格N-摘要。"""
    raw = spec.replace("×", "x").replace("*", "x")
    # 尽量取型号段
    token = raw.split()[0] if raw.split() else raw
    token = token[:12]
    return f"规格{idx}-{token}"



def build_report_spec_xlsx(result: dict[str, Any]) -> bytes:
    """导出精简多表 Excel：总说明 + 相对改动 + 按规格分表（不堆汇总）。"""
    wb = Workbook()
    used_titles: set[str] = set()
    specs = _ordered_specs(result)
    all_changes = list(result.get("changes") or [])
    all_tests = list(result.get("test_items") or [])
    if not specs:
        specs = ["目标规格"]

    # 1) 使用说明（短）
    ws_help = wb.active
    ws_help.title = _safe_sheet_title("使用说明", used_titles)
    ws_help["A1"] = "检验项目/修改说明参考包（AI，请复核）"
    _paint(ws_help["A1"], FILL_TITLE, FONT_TITLE)
    ws_help.merge_cells("A1:B1")
    help_rows = [
        ("看哪里", "先看「怎么改-总说明」和「相对原模版改了什么」，再打开对应规格分表。"),
        ("修改说明", "每规格一张：位置/原文/建议改/是否必须。红底=必须，绿底=建议。"),
        ("检验项目表", "每规格一张，列同报告第2页：序号/项目/单位/技术要求/结果草稿/评定；黄底为可誊写参考数据。"),
        ("重要", "正式报告必须换实测值；标准号/厚度/电阻以企标国标复核为准。"),
    ]
    _paint(ws_help.cell(3, 1, "项目"), FILL_HEADER, FONT_HEADER)
    _paint(ws_help.cell(3, 2, "说明"), FILL_HEADER, FONT_HEADER)
    for i, (a, b) in enumerate(help_rows, 4):
        fill = FILL_WARN if a == "重要" else FILL_LABEL
        _paint(ws_help.cell(i, 1, a), fill, FONT_BOLD if a == "重要" else None)
        _paint(
            ws_help.cell(i, 2, b),
            FILL_WARN if a == "重要" else None,
            FONT_WARN if a == "重要" else None,
        )
    ws_help.column_dimensions["A"].width = 12
    ws_help.column_dimensions["B"].width = 70

    # 2) 总说明
    ws0 = wb.create_sheet(_safe_sheet_title("怎么改-总说明", used_titles))
    ws0["A1"] = "怎么改（总说明）"
    _paint(ws0["A1"], FILL_TITLE, FONT_TITLE)
    ws0.merge_cells("A1:D1")
    _paint(ws0.cell(3, 1, "总览"), FILL_LABEL, FONT_BOLD)
    _paint(ws0.cell(3, 2, result.get("summary") or ""))
    _paint(ws0.cell(4, 1, "重要提醒"), FILL_WARN, FONT_WARN)
    warn = result.get("warnings") or []
    _paint(
        ws0.cell(
            4,
            2,
            "；".join(warn)
            if warn
            else "结果草稿非正式实测；材料/标准变化时勿照搬原指标。",
        ),
        FILL_WARN,
    )
    _paint(ws0.cell(6, 1, "建议套用哪一份样例"), FILL_TITLE, FONT_BOLD)
    ws0.merge_cells("A6:D6")
    for c, h in enumerate(["目标规格", "最接近样例", "样例原规格", "原因"], 1):
        _paint(ws0.cell(7, c, h), FILL_HEADER, FONT_HEADER)
    for i, m in enumerate(result.get("matches") or [], 8):
        vals = [
            m.get("target_spec") or "",
            m.get("base_report_no") or "",
            m.get("base_spec") or "",
            m.get("reason") or "",
        ]
        for c, v in enumerate(vals, 1):
            _paint(ws0.cell(i, c, v), FILL_ALT if i % 2 == 0 else None)
    for col, w in enumerate([28, 18, 28, 36], 1):
        ws0.column_dimensions[get_column_letter(col)].width = w

    # 3) 相对原模版改了什么
    ws_diff = wb.create_sheet(_safe_sheet_title("相对原模版改了什么", used_titles))
    ws_diff["A1"] = "相对源文件样例，本表关键改动（少而准）"
    _paint(ws_diff["A1"], FILL_TITLE, FONT_TITLE)
    ws_diff.merge_cells("A1:E1")
    for c, h in enumerate(["规格", "改动点", "原样例", "现在怎么改", "原因"], 1):
        _paint(ws_diff.cell(3, c, h), FILL_HEADER, FONT_HEADER)
    diffs = list(result.get("relative_diffs") or [])
    if not diffs:
        for c in all_changes:
            if not _is_must(c.get("must_change") or ""):
                continue
            diffs.append(
                {
                    "target_spec": c.get("target_spec") or "",
                    "aspect": c.get("position") or "",
                    "old_value": c.get("old_value") or "",
                    "new_value": c.get("new_value") or "",
                    "reason": c.get("note") or "",
                }
            )
        diffs = diffs[:16]
    for i, d in enumerate(diffs, 4):
        vals = [
            d.get("target_spec") or "",
            d.get("aspect") or "",
            d.get("old_value") or "",
            d.get("new_value") or "",
            d.get("reason") or "",
        ]
        for c, v in enumerate(vals, 1):
            fill = FILL_NEW if c == 4 else (FILL_ALT if i % 2 == 0 else None)
            _paint(ws_diff.cell(i, c, v), fill)
    for col, w in enumerate([26, 16, 28, 28, 28], 1):
        ws_diff.column_dimensions[get_column_letter(col)].width = w

    # 4) 按规格分表
    for idx, spec in enumerate(specs, 1):
        label = _short_spec_label(spec, idx)
        ch_rows = [c for c in all_changes if _s(c.get("target_spec")) == spec]
        te_rows = [t for t in all_tests if _s(t.get("target_spec")) == spec]
        if len(specs) == 1:
            if not ch_rows:
                ch_rows = all_changes
            if not te_rows:
                te_rows = all_tests

        ws_ch = wb.create_sheet(_safe_sheet_title(f"{label}-修改", used_titles))
        ws_ch["A1"] = f"规格{idx} 修改说明：{spec}"
        _paint(ws_ch["A1"], FILL_TITLE, FONT_TITLE)
        ws_ch.merge_cells("A1:E1")
        _write_sheet(
            ws_ch,
            ["位置", "原样例内容", "建议改为", "是否必须改", "备注"],
            [
                [
                    c.get("position") or "",
                    c.get("old_value") or "",
                    c.get("new_value") or "",
                    c.get("must_change") or "",
                    c.get("note") or "",
                ]
                for c in ch_rows
            ],
            start_row=3,
            must_col=4,
            new_col=3,
        )

        ws_te = wb.create_sheet(_safe_sheet_title(f"{label}-检验", used_titles))
        ws_te["A1"] = f"试样型号和规格：{spec}"
        _paint(ws_te["A1"], FILL_TITLE, FONT_TITLE)
        ws_te.merge_cells("A1:F1")
        ws_te["A2"] = "【检验结果列为示例草稿，正式报告请替换为实测值】"
        _paint(ws_te["A2"], FILL_WARN, FONT_WARN)
        ws_te.merge_cells("A2:F2")
        _write_sheet(
            ws_te,
            ["序号", "检验项目", "单位", "技术要求", "检验结果（示例草稿）", "单项评定"],
            [
                [
                    t.get("seq") or "",
                    t.get("item") or "",
                    t.get("unit") or "/",
                    t.get("requirement") or "",
                    t.get("result_draft") or "",
                    t.get("rating") or "",
                ]
                for t in te_rows
            ],
            start_row=4,
            draft_col=5,
            rating_col=6,
        )
        # 与样例一致的填写说明
        tip_row = 4 + 1 + max(len(te_rows), 0) + 1
        tip = (
            "填写说明：1）「检验结果（示例草稿）」按报告可誊写格式给出参考数据/结论，正式出报告前全部换成实测；"
            "2）单项评定：P符合，F不符合，N不要求判定，/本项无；"
            "3）技术要求若与企标不一致，以企标为准。"
        )
        cell_tip = ws_te.cell(tip_row, 1, tip)
        _paint(cell_tip, FILL_WARN, FONT_WARN)
        ws_te.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=6)

    # 5) 关键参数
    ws3 = wb.create_sheet(_safe_sheet_title("关键参数参考", used_titles))
    ws3["A1"] = "关键参数参考（正式以现行标准原文为准）"
    _paint(ws3["A1"], FILL_TITLE, FONT_TITLE)
    ws3.merge_cells("A1:D1")
    _write_sheet(
        ws3,
        ["目标规格", "项目", "常用参考值", "说明"],
        [
            [
                k.get("target_spec") or "",
                k.get("param") or "",
                k.get("ref_value") or "",
                k.get("note") or "",
            ]
            for k in (result.get("key_params") or [])
        ],
        start_row=3,
        param_col=3,
    )

    # 6) 操作步骤
    ws4 = wb.create_sheet(_safe_sheet_title("操作步骤清单", used_titles))
    ws4["A1"] = "实操步骤（按顺序做）"
    _paint(ws4["A1"], FILL_TITLE, FONT_TITLE)
    for i, step in enumerate(result.get("steps") or [], 3):
        _paint(ws4.cell(i, 1, step), FILL_STEP if i % 2 == 0 else FILL_ALT)
    ws4.column_dimensions["A"].width = 80

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
